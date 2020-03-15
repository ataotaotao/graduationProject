#-*-coding:utf-8-*-
'''
Created on 2018年10月20日

@author: DELL
'''
#-*-coding:utf-8-*-
'''
Created on 2018年8月15日

@author: DELL
'''

import numpy as np
from scipy import linalg
from scipy.io import loadmat
import matplotlib.pyplot as plt
from scipy.io import loadmat,savemat
from sklearn.cross_validation import train_test_split

from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

class Excel_format():
    def __init__(self, fileName):
        self.fileName = fileName
        
        
    def write_excel(self, master_name_o, result_dict):
        
        workbook = load_workbook(self.fileName)
        booksheet = workbook.get_sheet_by_name(master_name_o)
        
        
        std_num_list = result_dict['std_num_list']
        RMSEC_list = result_dict['RMSEC_list']
        RMSEP_list = result_dict['RMSEP_list' ]
        
        
        
        booksheet.cell(row = 1, column = 1, value = 'std_num')
        booksheet.cell(row = 1, column = 2, value = 'RMSEC')
        booksheet.cell(row = 1, column = 3, value = 'RMSEP')
        
        line = 2
        for j in range(len(std_num_list)):
            
            booksheet.cell(row = line, column = 1, value = std_num_list[j])
            booksheet.cell(row = line, column = 2, value = RMSEC_list[j])
            booksheet.cell(row = line, column = 3, value = RMSEP_list[j])
            line = line +1
        workbook.save(self.fileName)
        print "write complete!"
        
    
    
    
