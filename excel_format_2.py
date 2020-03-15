#-*-coding:utf-8-*-
'''
Created on 2018年10月20日

@author: DELL
'''
#-*-coding:utf-8-*-
'''
Created on 2018年8月15日

@author: DELL
'''

import numpy as np
from scipy import linalg
from scipy.io import loadmat
import matplotlib.pyplot as plt
from scipy.io import loadmat,savemat
from sklearn.cross_validation import train_test_split

from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

class Excel_format():
    def __init__(self, fileName):
        self.fileName = fileName
        
        
    def write_excel(self, master_name_o, result_dict , method_exclude):
        
        workbook = load_workbook(self.fileName)
        booksheet = workbook.get_sheet_by_name(master_name_o)
        
        change_rate_list = result_dict["change_rate_list"]
        p_value_list = result_dict['p_value_list']
        method_list = result_dict['method_list']
        std_list = result_dict['std_list']
        RMSEC_list = result_dict['RMSEC_list']
        RMSEP_list = result_dict['RMSEP_list' ]
        y_predict_list = result_dict['y_predict_list']
        comp_best_list = result_dict['comp_best_list']
        ycal_predict_list = result_dict['ycal_predict_list']
        
        
        num_method = len(method_list)
        num_std = len(std_list)
        print "size =", num_method, num_std 
        
        booksheet.cell(row = 1, column = 1, value = 'Methods_')
        booksheet.cell(row = 1, column = 2, value = num_method)
        booksheet.cell(row = 1, column = 4, value = 'Number')
        booksheet.cell(row = 1, column = 5, value = num_std)
        
        line = 2
        for j in range(num_std):
            booksheet.cell(row = line, column = 1, value = 'Number')
            booksheet.cell(row = line, column = 2, value = std_list[j])
            line = line + 1
            booksheet.cell(row = line, column = 2, value = 'LVs')
            booksheet.cell(row = line, column = 3, value = 'RMSEC')
            booksheet.cell(row = line, column = 4, value = 'RMSEP')
            booksheet.cell(row = line, column = 5, value = 'improve_rate')
            booksheet.cell(row = line, column = 6, value = 'wilcoxon')
            line = line + 1
            for i in range(num_method):
#                 print num_method
#                 print comp_best_list , j * num_method + i , j*num_method
                booksheet.cell(row = line+i, column = 1, value = method_list[i])
                booksheet.cell(row = line+i, column = 2, value = comp_best_list[j * num_method + i])
                booksheet.cell(row = line+i, column = 3, value = RMSEC_list[j * num_method + i])
                booksheet.cell(row = line+i, column = 4, value = RMSEP_list[j * num_method + i])
                if i >= method_exclude :
                    booksheet.cell(row = line+i, column = 5, value = change_rate_list[j * num_method + i - method_exclude])
                    booksheet.cell(row = line+i, column = 6, value = p_value_list[j * num_method + i - method_exclude])
                
            line = line + num_method
            booksheet.cell(row = line, column = 1, value = 'calibration')
            booksheet.cell(row = line, column = 2, value = 'y_predict')
            line = line + 1
            for i in range(num_method + 1):
                for k in range(len(ycal_predict_list[0])):
                    booksheet.cell(row = line+i, column = k+1, value = ycal_predict_list[j * (num_method + 1) + i][k][0]) 
#                 booksheet.append(y_predict_list[j * (num_method + 1) + i].ravel().tolist())
            line = line + num_method + 1 
            booksheet.cell(row = line, column = 1, value = 'test')
            booksheet.cell(row = line, column = 2, value = 'y_predict')
            line = line + 1
            for i in range(num_method + 1):
                for k in range(len(y_predict_list[0])):
                    booksheet.cell(row = line+i, column = k+1, value = y_predict_list[j * (num_method + 1) + i][k][0]) 
#                 booksheet.append(y_predict_list[j * (num_method + 1) + i].ravel().tolist())
            line = line + num_method + 2 
                   
        workbook.save(self.fileName)
        print "write complete!"
        
    
    
    
